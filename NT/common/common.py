import os
import openpyxl
import threading
from datetime import datetime


class Common(object):
    """公用函数"""
    api_cases_dict = {}  # api的所有用例dict
    api_cases_path = ""  # api用例路径
    time_format = "%Y%m%d%H%M%S"  # 时间格式
    start_time = str(datetime.now().strftime(time_format))  # 用例开始执行时间作文件夹名
    _instance_lock = threading.Lock()  # 设置单例锁

    def __new__(cls, *args, **kwargs):
        """单例模式(支持多线程)"""
        if not hasattr(cls, "_instance"):
            with cls._instance_lock:
                if not hasattr(cls, "_instance"):
                    cls._instance = object.__new__(cls)
        return cls._instance

    def get_now_time(self):
        """获取当前时间"""
        return datetime.now().strftime(self.time_format)

    def interval(self, start_time, end_time):
        """计算时间差"""
        return datetime.strptime(end_time, self.time_format) - datetime.strptime(start_time, self.time_format)

    @staticmethod
    def get_path(*args):
        """生成路径，参数为文件夹名或文件名，无参时为项目根路径"""
        try:
            # 获取当前目录的父目录路径(项目根路径)
            path = str(os.path.dirname(os.path.dirname(__file__)))

            # 循环生成路径
            for value in args:
                path = os.path.join(path, value).replace("\\", "/")  # 拼接路径并将"\"替换为"/",value为文件名时则自动创建
                # 不存在则创建文件夹
                if not os.path.exists(path) and "." not in value:  # 排出文件名并判断路径是否存在，不存在则创建文件夹
                    os.mkdir(path)
            return path
        except Exception as e:
            raise Exception("函数Common.get_path异常 %s" % e)

    def get_result_path(self, *args):
        """result目录下的日志，报告，截图路径"""
        # 自动生成result父目录和时间子目录，以及时间目录下的各级子目录
        result_path = Common.get_path("result", self.start_time, *args)
        return result_path

    def get_api_cases(self):
        """从Excel中读取cases"""
        try:
            # 拼接接口用例api_cases.xlsx路径,文件后缀只能是".xlsx"
            self.api_cases_path = Common.get_path("data", "api_cases.xlsx")

            # 打开xls文件
            wb = openpyxl.load_workbook(self.api_cases_path)
            # 获取workbook中所有的sheet
            sheets = wb.sheetnames

            # api_case_dict结构： case_dict = {case_key:case_value}/case_dict = {case_key:{param_key: param_value}}
            # case_value = {param_key: param_value}
            # api_cases_dict = {}  # case字典
            # origin = ""  # api用例请求地址原点
            api_sheet_dict = {}  # 每个sheet中的api用例
            case_key = []  # case_dict的key(case_name)
            case_value = {}  # case_dict的value(principal,remark,method,url,params,msg,code及其值组成的dict)
            case_param_key = []  # case_value的key(principal,remark,method,url,params,msg,code)
            # case_param_value = ""  # case_value的value(principal,remark,method,url,params,msg,code的值)

            print("Sheet数量：%s" % len(sheets))
            # 循环遍历所有sheet
            for i in range(len(sheets)):
                sheet = wb[sheets[i]]
                print("正在读取：%s" % sheet.title)

                origin = sheet.cell(row=1, column=1).value  # 获取第一行第一列的api用例请求地址原点
                if len(case_param_key) == 0:  # 判断是否已经添加case_param_key
                    for c in range(2, sheet.max_column + 1):  # 从第二列开始遍历每一列
                        param_key = sheet.cell(row=2, column=c).value  # 遍历第二行的所有列，获取case_param_key
                        case_param_key.append(param_key)

                for r in range(3, sheet.max_row + 1):  # 从第三行开始循环遍历所有行获取用例
                    if sheet.cell(row=r, column=1).value is not None:  # 判断每一行的第一列是否为空
                        key = sheet.cell(row=r, column=1).value  # 从第三行开始的第一列为用例名，用list case_key保存
                        if key not in self.api_cases_dict:
                            case_key.append(key)
                        else:
                            raise Exception("api用例名%s重复" % key)

                        for c in range(2, sheet.max_column + 1):  # 从第三行的第二列开始为用例数据，用字符串 case_param_value保存
                            case_param_value = sheet.cell(row=r, column=c).value  # 获取用例值
                            # 将参数类型case_param_key和用例数据组case_param_value成字典case_value
                            case_value[case_param_key[c - 2]] = case_param_value
                        # 将第一列的用例名case_key和用例数据字典case_value组成用例字典case_dict
                        api_sheet_dict[case_key[r - 3]] = case_value
                        case_value = {}  # 一个用例遍历完后用例值dict case_value置空
                    else:
                        break

                self.api_cases_dict[origin] = api_sheet_dict  # 将每一个sheet中的用例存入api_cases_dict
                case_key = []  # 一个sheet遍历完后用例名list case_key置空
                api_sheet_dict = {}

            return self.api_cases_path, self.api_cases_dict
        except Exception as e:
            raise Exception("Common.get_api_cases异常 %s" % e)
